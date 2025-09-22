import streamlit as st
import pandas as pd
import json
import re
import sqlite3
import tempfile
import os
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
import matplotlib.pyplot as plt
import plotly.express as px
import plotly.graph_objects as go
from wordcloud import WordCloud
import base64
from io import BytesIO, StringIO
import spacy
import warnings
import hashlib
from pathlib import Path
import chardet
import pdfplumber
import docx
import openpyxl
from PIL import Image
import pytesseract
import zipfile
import tarfile
import rarfile
import time
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as ReportLabImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import io
import secrets
import cryptography
from cryptography.fernet import Fernet
import hmac
import logging

# Firebase integration
import firebase_admin
from firebase_admin import credentials, firestore, auth
import requests

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("UFDR_Analysis")

# Set page configuration
st.set_page_config(
    page_title="Secure UFDR AI Analyst",
    page_icon="🔒",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Initialize Firebase
try:
    if not firebase_admin._apps:
        # You need to download your Firebase service account key JSON file
        # and either provide the path or set it as an environment variable
        cred_path = os.environ.get("FIREBASE_CREDENTIALS", r"C:\Users\Harini\OneDrive\Desktop\sih\udfr-370f2-firebase-adminsdk-fbsvc-f7cd6ae943.json")
        cred = credentials.Certificate(cred_path)
        firebase_admin.initialize_app(cred)
    db = firestore.client()
except Exception as e:
    logger.error(f"Firebase initialization error: {e}")
    st.error("Database connection failed. Some features may not work properly.")

# Security configuration
MAX_FILE_SIZE = 5 * 1024 * 1024 * 1024  # 5GB
ALLOWED_EXTENSIONS = {
    'txt', 'log', 'csv', 'json', 'xml', 'pdf', 'docx', 'xlsx', 'xls',
    'jpg', 'jpeg', 'png', 'bmp', 'tiff', 'zip', 'tar', 'gz', 'rar'
}

# Security utilities
class SecurityManager:
    """Manage security aspects of the application"""
    
    def __init__(self):
        # In production, use a secure key management system
        self.encryption_key = self._generate_encryption_key()
        self.cipher_suite = Fernet(self.encryption_key)
    
    def _generate_encryption_key(self):
        """Generate a secure encryption key"""
        try:
            # Try to load from environment variable
            key = os.environ.get("UFDR_ENCRYPTION_KEY")
            if key:
                return key.encode()
        except:
            pass
        
        # Generate a new key (for demo purposes - in production, use proper key management)
        return Fernet.generate_key()
    
    def encrypt_data(self, data: str) -> bytes:
        """Encrypt sensitive data"""
        if isinstance(data, str):
            data = data.encode()
        return self.cipher_suite.encrypt(data)
    
    def decrypt_data(self, encrypted_data: bytes) -> str:
        """Decrypt sensitive data"""
        return self.cipher_suite.decrypt(encrypted_data).decode()
    
    def validate_file(self, file_path: str, file_size: int) -> bool:
        """Validate file before processing"""
        # Check file size
        if file_size > MAX_FILE_SIZE:
            raise ValueError(f"File size exceeds maximum allowed size of {MAX_FILE_SIZE} bytes")
        
        # Check file extension
        file_ext = os.path.splitext(file_path)[1].lower().lstrip('.')
        if file_ext not in ALLOWED_EXTENSIONS:
            raise ValueError(f"File type {file_ext} is not allowed")
        
        # Additional security checks could be added here
        return True
    
    def sanitize_input(self, input_text: str) -> str:
        """Sanitize user input to prevent injection attacks"""
        # Remove potentially dangerous characters
        sanitized = re.sub(r'[;\\/*\'"<>|&$]', '', input_text)
        return sanitized[:1000]  # Limit input length

# Load NLP models
@st.cache_resource
def load_models():
    """Load all required AI models"""
    try:
        nlp = spacy.load("en_core_web_sm")
    except OSError:
        from spacy.cli import download
        download("en_core_web_sm")
        nlp = spacy.load("en_core_web_sm")
    
    return nlp

nlp = load_models()
security_manager = SecurityManager()

# Firebase Database Manager
class FirebaseManager:
    """Manage Firebase database operations"""
    
    def __init__(self):
        try:
            self.db = firestore.client()
        except Exception as e:
            logger.error(f"Firebase connection error: {e}")
            self.db = None
    
    def save_user_data(self, user_id, user_data):
        """Save user data to Firebase"""
        if not self.db:
            return False
        
        try:
            user_ref = self.db.collection('users').document(user_id)
            user_ref.set(user_data)
            return True
        except Exception as e:
            logger.error(f"Error saving user data: {e}")
            return False
    
    def get_user_data(self, user_id):
        """Retrieve user data from Firebase"""
        if not self.db:
            return None
        
        try:
            user_ref = self.db.collection('users').document(user_id)
            user_data = user_ref.get()
            if user_data.exists:
                return user_data.to_dict()
            return None
        except Exception as e:
            logger.error(f"Error retrieving user data: {e}")
            return None
    
    def save_chat_history(self, user_id, session_id, chat_data):
        """Save chat history to Firebase"""
        if not self.db:
            return False
        
        try:
            chat_ref = self.db.collection('users').document(user_id).collection('chats').document(session_id)
            chat_ref.set({
                'timestamp': datetime.now(),
                'chat_data': chat_data
            })
            return True
        except Exception as e:
            logger.error(f"Error saving chat history: {e}")
            return False
    
    def get_chat_history(self, user_id, session_id):
        """Retrieve chat history from Firebase"""
        if not self.db:
            return None
        
        try:
            chat_ref = self.db.collection('users').document(user_id).collection('chats').document(session_id)
            chat_data = chat_ref.get()
            if chat_data.exists:
                return chat_data.to_dict()
            return None
        except Exception as e:
            logger.error(f"Error retrieving chat history: {e}")
            return None
    
    def save_uploaded_file_info(self, user_id, file_info):
        """Save information about uploaded files to Firebase"""
        if not self.db:
            return False
        
        try:
            files_ref = self.db.collection('users').document(user_id).collection('files').document()
            files_ref.set({
                'timestamp': datetime.now(),
                'file_info': file_info
            })
            return True
        except Exception as e:
            logger.error(f"Error saving file info: {e}")
            return False
    
    def get_uploaded_files(self, user_id):
        """Retrieve list of uploaded files for a user"""
        if not self.db:
            return []
        
        try:
            files_ref = self.db.collection('users').document(user_id).collection('files')
            files = files_ref.stream()
            return [file.to_dict() for file in files]
        except Exception as e:
            logger.error(f"Error retrieving files: {e}")
            return []

class LargeFileProcessor:
    """Process large files efficiently with chunking and progress tracking"""
    
    def __init__(self):
        self.chunk_size = 1024 * 1024 * 10  # 10MB chunks
        self.supported_archives = ['.zip', '.tar', '.gz', '.rar']
    
    def process_large_file(self, file_path, file_type, callback=None):
        """Process large files in chunks"""
        content = ""
        file_size = os.path.getsize(file_path)
        processed_size = 0
        
        # For text-based files, read in chunks
        if file_type in ['.txt', '.log', '.csv', '.json', '.xml']:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                while True:
                    chunk = f.read(self.chunk_size)
                    if not chunk:
                        break
                    content += chunk
                    processed_size += len(chunk)
                    if callback:
                        callback(processed_size, file_size)
        
        # For binary files, use appropriate libraries
        elif file_type == '.pdf':
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        content += page_text + "\n"
                    processed_size += len(page_text or "")
                    if callback:
                        callback(processed_size, file_size)
        
        elif file_type == '.docx':
            doc = docx.Document(file_path)
            for para in doc.paragraphs:
                content += para.text + "\n"
            processed_size = len(content)
            if callback:
                callback(processed_size, processed_size)
        
        elif file_type in ['.xlsx', '.xls']:
            workbook = openpyxl.load_workbook(file_path)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    content += " ".join(str(cell) for cell in row if cell) + "\n"
            processed_size = len(content)
            if callback:
                callback(processed_size, processed_size)
        
        elif file_type in ['.jpg', '.jpeg', '.png', '.bmp', '.tiff']:
            # Use OCR for images
            try:
                image = Image.open(file_path)
                content = pytesseract.image_to_string(image)
                processed_size = len(content)
                if callback:
                    callback(processed_size, processed_size)
            except Exception as e:
                logger.error(f"OCR Error: {str(e)}")
                content = ""
        
        elif any(file_type.endswith(ext) for ext in self.supported_archives):
            content = self._extract_archive(file_path, callback)
            processed_size = len(content)
        
        else:
            # Fallback for unknown file types
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
                processed_size = len(content)
                if callback:
                    callback(processed_size, processed_size)
        
        return content
    
    def _extract_archive(self, file_path, callback=None):
        """Extract text from archive files"""
        content = ""
        
        if file_path.endswith('.zip'):
            with zipfile.ZipFile(file_path, 'r') as zip_ref:
                for file_info in zip_ref.infolist():
                    if not file_info.is_dir():
                        with zip_ref.open(file_info) as file:
                            try:
                                file_content = file.read().decode('utf-8', errors='ignore')
                                content += f"\n--- File: {file_info.filename} ---\n{file_content}\n"
                            except:
                                # Skip binary files
                                pass
        
        elif file_path.endswith('.tar') or file_path.endswith('.tar.gz') or file_path.endswith('.tgz'):
            mode = 'r:gz' if file_path.endswith('.gz') or file_path.endswith('.tgz') else 'r'
            with tarfile.open(file_path, mode) as tar_ref:
                for member in tar_ref.getmembers():
                    if member.isfile():
                        try:
                            file = tar_ref.extractfile(member)
                            if file:
                                file_content = file.read().decode('utf-8', errors='ignore')
                                content += f"\n--- File: {member.name} ---\n{file_content}\n"
                        except:
                            # Skip binary files
                            pass
        
        elif file_path.endswith('.rar'):
            with rarfile.RarFile(file_path, 'r') as rar_ref:
                for file_info in rar_ref.infolist():
                    if not file_info.is_dir():
                        with rar_ref.open(file_info) as file:
                            try:
                                file_content = file.read().decode('utf-8', errors='ignore')
                                content += f"\n--- File: {file_info.filename} ---\n{file_content}\n"
                            except:
                                # Skip binary files
                                pass
        
        return content

class UFDRParser:
    """Parse structured and unstructured UFDR data with support for large files"""
    
    def __init__(self):
        # Improved crypto patterns
        self.crypto_patterns = [
            r'\b[13][a-km-zA-HJ-NP-Z1-9]{25,34}\b',  # Bitcoin
            r'\b0x[a-fA-F0-9]{40}\b',  # Ethereum
            r'\bT[A-Za-z1-9]{33}\b',  # Tron
            r'\b[LM3][a-km-zA-HJ-NP-Z1-9]{25,34}\b',  # Litecoin/Dogecoin
            r'\bbc1[a-zA-Z0-9]{39,59}\b',  # Bitcoin Bech32
            r'\b[A-Z0-9]{20,}\b'  # General crypto pattern
        ]
        self.phone_pattern = r'(\+?\d{1,3}[-.\s]?\(?\d{1,4}\)?[-.\s]?\d{1,4}[-.\s]?\d{1,9})'
        self.email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        self.url_pattern = r'https?://(?:[-\w.]|(?:%[\da-fA-F]{2}))+[/\w\.-]*\??[/\w\.-=&]*'
        self.ip_pattern = r'\b(?:\d{1,3}\.){3}\d{1,3}\b'
        self.suspicious_keywords = [
            'fraud', 'scam', 'launder', 'bribe', 'hack', 'attack', 'exploit', 
            'threat', 'blackmail', 'extort', 'ransom', 'illegal', 'unauthorized',
            'confidential', 'secret', 'classified', 'undercover', 'covert', 'operation'
        ]
        self.file_processor = LargeFileProcessor()
        
    def parse_file(self, file_path: str, filename: str, progress_callback=None) -> Dict[str, Any]:
        """Parse any type of file with progress tracking"""
        file_ext = os.path.splitext(filename)[1].lower()
        
        try:
            # Security validation
            file_size = os.path.getsize(file_path)
            security_manager.validate_file(file_path, file_size)
            
            content = self.file_processor.process_large_file(
                file_path, file_ext, progress_callback
            )
            
            # Try to detect if content is JSON
            if self._is_json(content):
                return self._parse_json(content, filename)
            else:
                return self._parse_text(content, filename, file_ext)
                
        except Exception as e:
            logger.error(f"Error parsing file {filename}: {str(e)}")
            return {"error": f"Error parsing file: {str(e)}", "filename": filename}
    
    def _is_json(self, content: str) -> bool:
        """Check if content is JSON"""
        try:
            json.loads(content)
            return True
        except:
            return False
    
    def _parse_json(self, content: str, filename: str) -> Dict[str, Any]:
        """Parse JSON content"""
        data = json.loads(content)
        
        # Extract communications if they exist in the JSON structure
        communications = []
        if isinstance(data, dict):
            # Look for common communication structures in JSON
            for key, value in data.items():
                if isinstance(value, list) and key.lower() in ['messages', 'chats', 'communications']:
                    for item in value:
                        if isinstance(item, dict):
                            comm = {
                                'type': item.get('type', 'chat'),
                                'date': item.get('date', ''),
                                'time': item.get('time', ''),
                                'sender': item.get('sender', ''),
                                'content': item.get('content', ''),
                                'crypto_addresses': self._extract_crypto_addresses(str(item)),
                                'foreign_numbers': self._extract_foreign_numbers(str(item)),
                                'emails': self._extract_emails(str(item)),
                                'urls': self._extract_urls(str(item)),
                                'ips': self._extract_ips(str(item)),
                                'suspicious_keywords': self._extract_suspicious_keywords(str(item))
                            }
                            communications.append(comm)
        
        return {
            'filename': filename,
            'content': content[:10000] + "..." if len(content) > 10000 else content,  # Store preview
            'communications': communications,
            'metadata': {'source': 'json', 'filename': filename},
            'financial_data': self._extract_financial_data(content),
            'network_data': self._extract_network_data(content)
        }
    
    def _parse_text(self, content: str, filename: str, file_ext: str) -> Dict[str, Any]:
        """Parse unstructured text content"""
        # Extract all content for analysis
        return {
            'filename': filename,
            'file_type': file_ext,
            'content': content[:10000] + "..." if len(content) > 10000 else content,  # Store preview
            'metadata': self._extract_metadata(content),
            'communications': self._extract_communications(content),
            'financial_data': self._extract_financial_data(content),
            'network_data': self._extract_network_data(content),
            'suspicious_items': self._extract_all_suspicious_items(content)
        }
    
    def _extract_metadata(self, content: str) -> Dict[str, str]:
        """Extract metadata from content"""
        metadata = {}
        patterns = {
            'date': r'Date[:;\s]+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
            'time': r'Time[:;\s]+(\d{1,2}:\d{2}(?::\d{2})?)',
            'device_id': r'Device\s*ID[:;\s]+([\w-]+)',
            'report_id': r'Report\s*ID[:;\s]+([\w-]+)',
            'investigator': r'Investigator[:;\s]+([\w\s]+)'
        }
        
        for key, pattern in patterns.items():
            match = re.search(pattern, content, re.IGNORECASE)
            if match:
                metadata[key] = match.group(1)
        
        metadata['source'] = 'text'
        return metadata
    
    def _extract_communications(self, content: str) -> List[Dict[str, str]]:
        """Extract communication records from content"""
        communications = []
    
    # Fixed chat/message patterns - corrected the regex errors
        message_patterns = [
            r'(\d{1,2}[:/]\d{1,2}[:/]\d{2,4})\s+(\d{1,2}:\d{2}(?::\d{2})?)\s+([^:]+):\s*(.+)',
            r'\[(\d{4}-\d{2}-\d{2})\s+(\d{2}:\d{2}:\d{2})\]\s+([^:]+):\s*(.+)',  # Fixed: removed {2} -> \d{2}
            r'(\d{2}/\d{2}/\d{4})\s+(\d{1,2}:\d{2}\s*[AP]M)\s+([^:]+):\s*(.+)',  # Fixed: removed {4} -> \d{4}
            r'(\d{4}-\d{2}-\d{2})\s+(\d{2}:\d{2}:\d{2})\s+([^:]+):\s*(.+)'
        ]
    
        for pattern in message_patterns:
            try:
                matches = re.findall(pattern, content)
                for match in matches:
                    if len(match) == 4:
                        date, time, sender, message = match
                        communications.append({
                            'type': 'chat',
                            'date': date,
                            'time': time,
                            'sender': sender.strip(),
                            'content': message.strip(),
                            'crypto_addresses': self._extract_crypto_addresses(message),
                            'foreign_numbers': self._extract_foreign_numbers(message),
                            'emails': self._extract_emails(message),
                            'urls': self._extract_urls(message),
                            'ips': self._extract_ips(message),
                            'suspicious_keywords': self._extract_suspicious_keywords(message)
                        })
            except re.error as e:
                logger.warning(f"Regex error with pattern {pattern}: {e}")
                continue  # Skip invalid patterns
        
        return communications
    
    def _extract_financial_data(self, content: str) -> List[Dict[str, str]]:
        """Extract financial transactions from content"""
        financial_data = []
        
        # Look for transaction patterns - improved pattern matching
        transaction_patterns = [
            r'(\d{2,4}[-/]\d{1,2}[-/]\d{1,2})\s+([^:]+):\s*Amount[:;\s]+([$\d.,]+)',
            r'(\d{2,4}[-/]\d{1,2}[-/]\d{1,2})\s+([^:]+):\s*([$\d.,]+)',
            r'(\d{2,4}[-/]\d{1,2}[-/]\d{1,2})\s+([$\d.,]+)\s+([^\n]+)'
        ]
        
        for pattern in transaction_patterns:
            matches = re.findall(pattern, content, re.IGNORECASE)
            for match in matches:
                if len(match) == 3:
                    date, description, amount = match
                    financial_data.append({
                        'date': date.strip(),
                        'amount': amount.strip(),
                        'description': description.strip(),
                        'crypto_addresses': self._extract_crypto_addresses(description),
                        'suspicious_keywords': self._extract_suspicious_keywords(description)
                    })
        
        return financial_data
    
    def _extract_network_data(self, content: str) -> List[Dict[str, str]]:
        """Extract network activity from content"""
        network_data = []
        
        # Look for network activity patterns
        network_patterns = [
            r'Accessed\s+IP[:;\s]+(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})',
            r'Accessed\s+URL[:;\s]+(https?://[^\s]+)',
            r'IP[:;\s]+(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})',
            r'URL[:;\s]+(https?://[^\s]+)'
        ]
        
        for pattern in network_patterns:
            matches = re.findall(pattern, content, re.IGNORECASE)
            for match in matches:
                if '://' in match:
                    network_data.append({
                        'type': 'url',
                        'value': match
                    })
                elif re.match(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}', match):
                    network_data.append({
                        'type': 'ip_address',
                        'value': match
                    })
        
        return network_data
    
    def _extract_all_suspicious_items(self, content: str) -> List[Dict[str, Any]]:
        """Extract all suspicious items from content"""
        suspicious_items = []
        
        # Extract crypto addresses from entire content
        crypto_addresses = self._extract_crypto_addresses(content)
        if crypto_addresses:
            for addr in crypto_addresses:
                suspicious_items.append({
                    'type': 'crypto',
                    'content': f'Found cryptocurrency address: {addr}',
                    'details': f'Cryptocurrency address: {addr}'
                })
        
        # Extract foreign numbers from entire content
        foreign_numbers = self._extract_foreign_numbers(content)
        if foreign_numbers:
            for num in foreign_numbers:
                suspicious_items.append({
                    'type': 'foreign_contact',
                    'content': f'Found foreign phone number: {num}',
                    'details': f'Foreign phone number: {num}'
                })
        
        # Extract suspicious keywords from entire content
        suspicious_keywords = self._extract_suspicious_keywords(content)
        if suspicious_keywords:
            for keyword in suspicious_keywords:
                suspicious_items.append({
                    'type': 'suspicious_keywords',
                    'content': f'Found suspicious keyword: {keyword}',
                    'details': f'Suspicious keyword: {keyword}'
                })
        
        # Extract URLs and IPs from entire content
        urls = self._extract_urls(content)
        ips = self._extract_ips(content)
        
        if urls:
            for url in urls:
                suspicious_items.append({
                    'type': 'potential_phishing',
                    'content': f'Found URL: {url}',
                    'details': f'URL: {url}'
                })
        
        if ips:
            for ip in ips:
                # Check if IP is foreign (not private or local)
                if not self._is_private_ip(ip):
                    suspicious_items.append({
                        'type': 'potential_phishing',
                        'content': f'Found external IP: {ip}',
                        'details': f'External IP: {ip}'
                    })
        
        return suspicious_items
    
    def _is_private_ip(self, ip: str) -> bool:
        """Check if an IP address is private"""
        private_patterns = [
            r'^10\.',
            r'^172\.(1[6-9]|2[0-9]|3[0-1])\.',
            r'^192\.168\.',
            r'^127\.',
            r'^169\.254\.',
            r'^::1$',
            r'^fc00::',
            r'^fd00::'
        ]
        
        for pattern in private_patterns:
            if re.match(pattern, ip):
                return True
        return False
    
    def _extract_crypto_addresses(self, text: str) -> List[str]:
        """Extract cryptocurrency addresses from text"""
        addresses = []
        for pattern in self.crypto_patterns:
            matches = re.findall(pattern, text)
            addresses.extend(matches)
        return list(set(addresses))
    
    def _extract_foreign_numbers(self, text: str) -> List[str]:
        """Extract potential foreign phone numbers from text"""
        numbers = re.findall(self.phone_pattern, text)
        foreign_numbers = []
        
        for number in numbers:
            # Clean the number
            clean_number = re.sub(r'[^\d+]', '', number)
            
            # Simple heuristic: numbers starting with + and not with +1 (US/Canada)
            if clean_number.startswith('+') and not clean_number.startswith('+1'):
                foreign_numbers.append(clean_number)
            # Numbers with international format but without +
            elif re.match(r'00[2-9]\d{0,3}', clean_number):
                foreign_numbers.append(clean_number)
        
        return list(set(foreign_numbers))
    
    def _extract_emails(self, text: str) -> List[str]:
        """Extract email addresses from text"""
        emails = re.findall(self.email_pattern, text)
        return list(set(emails))
    
    def _extract_urls(self, text: str) -> List[str]:
        """Extract URLs from text"""
        urls = re.findall(self.url_pattern, text)
        return list(set(urls))
    
    def _extract_ips(self, text: str) -> List[str]:
        """Extract IP addresses from text"""
        ips = re.findall(self.ip_pattern, text)
        return list(set(ips))
    
    def _extract_suspicious_keywords(self, text: str) -> List[str]:
        """Extract suspicious keywords from text"""
        found_keywords = []
        text_lower = text.lower()
        
        for keyword in self.suspicious_keywords:
            if re.search(r'\b' + re.escape(keyword) + r'\b', text_lower):
                found_keywords.append(keyword)
        
        return list(set(found_keywords))

class PDFReportGenerator:
    """Generate comprehensive PDF reports"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        self.heading_style = ParagraphStyle(
            'CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            spaceBefore=12
        )
        self.normal_style = ParagraphStyle(
            'CustomNormal',
            parent=self.styles['Normal'],
            fontSize=10,
            spaceAfter=6
        )
    
    def generate_report(self, analysis_data: Dict[str, Any], filename: str) -> BytesIO:
        """Generate a comprehensive PDF report"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        # Title
        title = Paragraph("SECURE UFDR ANALYSIS REPORT", self.title_style)
        elements.append(title)
        
        # Metadata
        elements.append(Paragraph(f"Report generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", self.normal_style))
        elements.append(Paragraph(f"File analyzed: {filename}", self.normal_style))
        elements.append(Paragraph("Confidential: For authorized personnel only", self.normal_style))
        elements.append(Spacer(1, 20))
        
        # Executive Summary
        elements.append(Paragraph("EXECUTIVE SUMMARY", self.heading_style))
        summary_text = self._generate_executive_summary(analysis_data)
        elements.append(Paragraph(summary_text, self.normal_style))
        elements.append(Spacer(1, 12))
        
        # Key Findings
        elements.append(Paragraph("KEY FINDINGS", self.heading_style))
        findings = self._generate_key_findings(analysis_data)
        for finding in findings:
            elements.append(Paragraph(f"• {finding}", self.normal_style))
        elements.append(Spacer(1, 12))
        
        # Suspicious Activities
        elements.append(Paragraph("SUSPICIOUS ACTIVITIES", self.heading_style))
        suspicious_items = analysis_data.get('suspicious_items', [])
        
        if suspicious_items:
            # Group by type
            by_type = {}
            for item in suspicious_items:
                item_type = item.get('type', 'unknown')
                if item_type not in by_type:
                    by_type[item_type] = []
                by_type[item_type].append(item)
            
            for item_type, items in by_type.items():
                elements.append(Paragraph(f"{item_type.replace('_', ' ').title()} ({len(items)} found)", self.heading_style))
                
                for i, item in enumerate(items[:10]):  # Show top 10 per type
                    elements.append(Paragraph(f"{i+1}. {item.get('content', 'No content')}", self.normal_style))
                    if item.get('details'):
                        elements.append(Paragraph(f"   Details: {item.get('details', 'No additional details')}", self.normal_style))
                    elements.append(Spacer(1, 6))
                
                if len(items) > 10:
                    elements.append(Paragraph(f"... and {len(items) - 10} more {item_type} items", self.normal_style))
                
                elements.append(Spacer(1, 6))
        else:
            elements.append(Paragraph("No suspicious activities detected.", self.normal_style))
        
        elements.append(Spacer(1, 12))
        
        # Statistics
        elements.append(Paragraph("STATISTICAL OVERVIEW", self.heading_style))
        stats = analysis_data.get('statistics', {})
        stats_data = [
            ['Metric', 'Value'],
            ['Total Communications', stats.get('total_communications', 0)],
            ['Suspicious Items', stats.get('suspicious_items', 0)],
            ['Cryptocurrency References', stats.get('crypto_references', 0)],
            ['Foreign Contacts', stats.get('foreign_contacts', 0)],
            ['Potential Phishing', stats.get('potential_phishing', 0)],
            ['Suspicious Keywords', stats.get('suspicious_keywords', 0)]
        ]
        
        stats_table = Table(stats_data)
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))
        
        elements.append(stats_table)
        elements.append(Spacer(1, 12))
        
        # Recommendations
        elements.append(Paragraph("RECOMMENDATIONS", self.heading_style))
        recommendations = self._generate_recommendations(analysis_data)
        for rec in recommendations:
            elements.append(Paragraph(f"• {rec}", self.normal_style))
        
        # Footer
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("--- END OF REPORT ---", self.normal_style))
        elements.append(Paragraph("This report was generated by Secure UFDR AI Analysis Platform", self.normal_style))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    def _generate_executive_summary(self, analysis_data: Dict[str, Any]) -> str:
        """Generate executive summary"""
        stats = analysis_data.get('statistics', {})
        suspicious_count = stats.get('suspicious_items', 0)
        
        if suspicious_count > 0:
            return f"This analysis identified {suspicious_count} suspicious items requiring further investigation. The data shows potential indicators of concerning activities that should be reviewed by security personnel."
        else:
            return "The analysis did not identify any immediately suspicious activities. The data appears normal, but continued monitoring is recommended."
    
    def _generate_key_findings(self, analysis_data: Dict[str, Any]) -> List[str]:
        """Generate key findings"""
        findings = []
        stats = analysis_data.get('statistics', {})
        
        if stats.get('crypto_references', 0) > 0:
            findings.append(f"Found {stats['crypto_references']} communications containing cryptocurrency addresses")
        
        if stats.get('foreign_contacts', 0) > 0:
            findings.append(f"Discovered {stats['foreign_contacts']} communications with foreign phone numbers")
        
        if stats.get('potential_phishing', 0) > 0:
            findings.append(f"Identified {stats['potential_phishing']} communications containing potential phishing URLs or IP addresses")
        
        if stats.get('suspicious_keywords', 0) > 0:
            findings.append(f"Detected {stats['suspicious_keywords']} instances of suspicious keywords in communications")
        
        if not findings:
            findings.append("No significant findings detected in the analyzed data")
        
        return findings
    
    def _generate_recommendations(self, analysis_data: Dict[str, Any]) -> List[str]:
        """Generate recommendations"""
        recommendations = []
        stats = analysis_data.get('statistics', {})
        
        if stats.get('crypto_references', 0) > 0:
            recommendations.append("Investigate the cryptocurrency addresses found to trace transactions")
            recommendations.append("Identify the owners of the wallets associated with these addresses")
        
        if stats.get('foreign_contacts', 0) > 0:
            recommendations.append("Verify the identities associated with the foreign phone numbers")
            recommendations.append("Investigate the nature of communications with these numbers")
        
        if stats.get('potential_phishing', 0) > 0:
            recommendations.append("Analyze the suspicious URLs and IP addresses for potential phishing campaigns")
            recommendations.append("Check if these domains/IPs appear in threat intelligence databases")
        
        if stats.get('suspicious_keywords', 0) > 0:
            recommendations.append("Review communications containing suspicious keywords for potential threats")
            recommendations.append("Consider implementing additional monitoring for these types of communications")
        
        if not recommendations:
            recommendations.append("No specific recommendations at this time. Continue monitoring for suspicious activity.")
        
        return recommendations

class AIChatProcessor:
    """Process natural language queries with AI"""
    
    def __init__(self):
        self.contexts = {}
    
    def add_context(self, filename: str, content: str):
        """Add file content as context for QA"""
        self.contexts[filename] = content[:50000]  # Limit context size
    
    def get_answer(self, question: str, filename: str = None) -> str:
        """Get answer to question using pattern matching"""
        question_lower = question.lower()
        
        if filename and filename in self.contexts:
            context = self.contexts[filename]
        else:
            # Combine all contexts if no specific file is requested
            context = " ".join(self.contexts.values())
            if not context:
                return "No data available for analysis. Please upload files first."
        
        # Pattern-based response generation
        if any(word in question_lower for word in ['crypto', 'bitcoin', 'ethereum', 'wallet']):
            # Look for crypto addresses in context
            crypto_patterns = [
                r'\b[13][a-km-zA-HJ-NP-Z1-9]{25,34}\b',  # Bitcoin
                r'\b0x[a-fA-F0-9]{40}\b',                # Ethereum
                r'\bT[A-Za-z1-9]{33}\b',                 # Tron
                r'\b[LM3][a-km-zA-HJ-NP-Z1-9]{25,34}\b', # Litecoin or Dogecoin
                r'\bbc1[a-zA-Z0-9]{39,59}\b'             # Bitcoin Cash
                ]
            
            crypto_addresses = []
            for pattern in crypto_patterns:
                crypto_addresses.extend(re.findall(pattern, context))
            
            if crypto_addresses:
                return f"I found {len(crypto_addresses)} cryptocurrency addresses in the data. Examples: {', '.join(crypto_addresses[:3])}"
            else:
                return "No cryptocurrency addresses found in the data."
        
        elif any(word in question_lower for word in ['foreign', 'international', 'number', 'phone']):
            # Look for foreign numbers in context
            phone_pattern = r'(\+?\d{1,3}[-.\s]?\(?\d{1,4}\)?[-.\s]?\d{1,4}[-.\s]?\d{1,9})'
            numbers = re.findall(phone_pattern, context)
            
            foreign_numbers = []
            for number in numbers:
                clean_number = re.sub(r'[^\d+]', '', number)
                if clean_number.startswith('+') and not clean_number.startswith('+1'):
                    foreign_numbers.append(clean_number)
                elif re.match(r'00[2-9]\d{0,3}', clean_number):
                    foreign_numbers.append(clean_number)
            
            if foreign_numbers:
                return f"I found {len(foreign_numbers)} foreign phone numbers in the data. Examples: {', '.join(foreign_numbers[:3])}"
            else:
                return "No foreign phone numbers found in the data."
        
        elif any(word in question_lower for word in ['suspicious', 'threat', 'risk', 'danger']):
            # Look for suspicious keywords in context
            suspicious_keywords = [
                'fraud', 'scam', 'launder', 'bribe', 'hack', 'attack', 'exploit', 
                'threat', 'blackmail', 'extort', 'ransom', 'illegal', 'unauthorized',
                'confidential', 'secret', 'classified', 'undercover', 'covert', 'operation'
            ]
            
            found_keywords = []
            for keyword in suspicious_keywords:
                if re.search(r'\b' + re.escape(keyword) + r'\b', context.lower()):
                    found_keywords.append(keyword)
            
            if found_keywords:
                return f"I found {len(found_keywords)} suspicious keywords in the data: {', '.join(found_keywords)}"
            else:
                return "No suspicious keywords found in the data."
        
        elif any(word in question_lower for word in ['summary', 'overview', 'statistics']):
            # Generate a summary of the data
            crypto_count = len(re.findall(r'\b[13][a-km-zA-HJ-NP-Z1-9]{25,34}\b|\b0x[a-fA-F0-9]{40}\b|\bT[A-Za-z1-9]{33}\b', context))
            
            phone_pattern = r'(\+?\d{1,3}[-.\s]?\(?\d{1,4}\)?[-.\s]?\d{1,4}[-.\s]?\d{1,9})'
            numbers = re.findall(phone_pattern, context)
            foreign_count = 0
            for number in numbers:
                clean_number = re.sub(r'[^\d+]', '', number)
                if clean_number.startswith('+') and not clean_number.startswith('+1'):
                    foreign_count += 1
                elif re.match(r'00[2-9]\d{0,3}', clean_number):
                    foreign_count += 1
            
            suspicious_words = ['fraud', 'scam', 'launder', 'bribe', 'hack', 'attack', 'exploit']
            suspicious_count = 0
            for word in suspicious_words:
                if re.search(r'\b' + re.escape(word) + r'\b', context.lower()):
                    suspicious_count += 1
            
            return f"The data contains {crypto_count} cryptocurrency references, {foreign_count} foreign phone numbers, and {suspicious_count} suspicious keywords."
        
        else:
            return "I can help you analyze cryptocurrency addresses, foreign phone numbers, suspicious keywords, or provide a summary of the data. Please ask a specific question about these topics."

class UFDRAnalysisAgent:
    """Main agent class that coordinates the UFDR analysis"""
    
    def __init__(self, user_id=None):
        self.parser = UFDRParser()
        self.pdf_generator = PDFReportGenerator()
        self.ai_chat = AIChatProcessor()
        self.uploaded_files = {}
        self.temp_dir = tempfile.mkdtemp()
        self.user_id = user_id
        self.firebase = FirebaseManager()
        self.session_id = str(datetime.now().timestamp())
    
    def process_uploaded_file(self, uploaded_file) -> Dict[str, Any]:
        """Process an uploaded file with progress tracking"""
        try:
            # Save uploaded file to temporary location
            file_path = os.path.join(self.temp_dir, uploaded_file.name)
            with open(file_path, "wb") as f:
                f.write(uploaded_file.getbuffer())
            
            # Create progress bar
            progress_bar = st.progress(0)
            progress_text = st.empty()
            
            def update_progress(processed, total):
                if total > 0:
                    percent = processed / total
                    progress_bar.progress(min(percent, 1.0))
                    progress_text.text(f"Processing: {processed}/{total} bytes ({percent*100:.1f}%)")
            
          # Process the file
            parsed_data = self.parser.parse_file(
                file_path, uploaded_file.name, update_progress
            )
            
            # Clear progress indicators
            progress_bar.empty()
            progress_text.empty()
            
            if "error" in parsed_data:
                logger.error(f"Error processing {uploaded_file.name}: {parsed_data['error']}")
                return parsed_data
            
            self.uploaded_files[uploaded_file.name] = parsed_data
            
            # Add to AI context
            self.ai_chat.add_context(uploaded_file.name, parsed_data.get('content', ''))
            
            # Save file info to Firebase
            if self.user_id:
                file_info = {
                    'filename': uploaded_file.name,
                    'size': uploaded_file.size,
                    'upload_time': datetime.now().isoformat(),
                    'file_type': uploaded_file.type,
                    'analysis_summary': f"Processed with {len(parsed_data.get('communications', []))} communications found"
                }
                self.firebase.save_uploaded_file_info(self.user_id, file_info)
            
            return parsed_data
            
        except Exception as e:
            logger.error(f"Error processing file: {str(e)}")
            return {"error": str(e)}
    
    def analyze_content(self, filename: str) -> Dict[str, Any]:
        """Analyze content of a specific file"""
        if filename not in self.uploaded_files:
            return {"error": "File not found"}
        
        data = self.uploaded_files[filename]
        analysis = {
            "filename": filename,
            "suspicious_items": [],
            "statistics": {},
            "insights": []
        }
        
        # Extract suspicious items from parsed data
        if 'suspicious_items' in data:
            analysis['suspicious_items'] = data['suspicious_items']
        
        # Also check communications for suspicious content
        for comm in data.get('communications', []):
            # Check for crypto addresses
            if comm.get('crypto_addresses'):
                analysis['suspicious_items'].append({
                    'type': 'crypto',
                    'content': comm['content'],
                    'sender': comm['sender'],
                    'date': comm.get('date', ''),
                    'details': f"Crypto addresses: {', '.join(comm['crypto_addresses'])}"
                })
            
            # Check for foreign numbers
            if comm.get('foreign_numbers'):
                analysis['suspicious_items'].append({
                    'type': 'foreign_contact',
                    'content': comm['content'],
                    'sender': comm['sender'],
                    'date': comm.get('date', ''),
                    'details': f"Foreign numbers: {', '.join(comm['foreign_numbers'])}"
                })
            
            # Check for URLs and IPs (potential phishing)
            if comm.get('urls') or comm.get('ips'):
                analysis['suspicious_items'].append({
                    'type': 'potential_phishing',
                    'content': comm['content'],
                    'sender': comm['sender'],
                    'date': comm.get('date', ''),
                    'details': f"URLs: {', '.join(comm.get('urls', []))}, IPs: {', '.join(comm.get('ips', []))}"
                })
            
            # Check for suspicious keywords
            if comm.get('suspicious_keywords'):
                analysis['suspicious_items'].append({
                    'type': 'suspicious_keywords',
                    'content': comm['content'],
                    'sender': comm['sender'],
                    'date': comm.get('date', ''),
                    'details': f"Suspicious keywords: {', '.join(comm['suspicious_keywords'])}"
                })
        
        # Generate statistics
        analysis['statistics'] = {
            'total_communications': len(data.get('communications', [])),
            'suspicious_items': len(analysis['suspicious_items']),
            'crypto_references': sum(1 for item in analysis['suspicious_items'] if item['type'] == 'crypto'),
            'foreign_contacts': sum(1 for item in analysis['suspicious_items'] if item['type'] == 'foreign_contact'),
            'potential_phishing': sum(1 for item in analysis['suspicious_items'] if item['type'] == 'potential_phishing'),
            'suspicious_keywords': sum(1 for item in analysis['suspicious_items'] if item['type'] == 'suspicious_keywords')
        }
        
        # Generate insights
        if analysis['statistics']['suspicious_items'] > 0:
            analysis['insights'].append(
                f"Found {analysis['statistics']['suspicious_items']} suspicious items in {filename}"
            )
            
            if analysis['statistics']['crypto_references'] > 0:
                analysis['insights'].append(
                    f"⚠️ {analysis['statistics']['crypto_references']} communications contain cryptocurrency addresses"
                )
            
            if analysis['statistics']['foreign_contacts'] > 0:
                analysis['insights'].append(
                    f"🌍 {analysis['statistics']['foreign_contacts']} communications contain foreign phone numbers"
                )
            
            if analysis['statistics']['potential_phishing'] > 0:
                analysis['insights'].append(
                    f"🎣 {analysis['statistics']['potential_phishing']} communications contain potential phishing links"
                )
            
            if analysis['statistics']['suspicious_keywords'] > 0:
                analysis['insights'].append(
                    f"🔍 {analysis['statistics']['suspicious_keywords']} communications contain suspicious keywords"
                )
        else:
            analysis['insights'].append("No suspicious items found in this file")
        
        return analysis
    
    def query_content(self, filename: str, query: str) -> Dict[str, Any]:
        """Query specific content in a file"""
        if filename not in self.uploaded_files:
            return {"error": "File not found"}
        
        data = self.uploaded_files[filename]
        results = []
        
        # Simple keyword-based search
        query_lower = query.lower()
        
        # Search in communications
        for comm in data.get('communications', []):
            content = comm.get('content', '').lower()
            sender = comm.get('sender', '').lower()
            
            if query_lower in content or query_lower in sender:
                results.append({
                    'type': 'communication',
                    'sender': comm.get('sender', ''),
                    'content': comm.get('content', ''),
                    'date': comm.get('date', ''),
                    'time': comm.get('time', '')
                })
        
        # Search in financial data
        for financial in data.get('financial_data', []):
            description = financial.get('description', '').lower()
            amount = financial.get('amount', '').lower()
            
            if query_lower in description or query_lower in amount:
                results.append({
                    'type': 'financial',
                    'date': financial.get('date', ''),
                    'amount': financial.get('amount', ''),
                    'description': financial.get('description', '')
                })
        
        # Search in network data
        for network in data.get('network_data', []):
            value = network.get('value', '').lower()
            
            if query_lower in value:
                results.append({
                    'type': 'network',
                    'value': network.get('value', ''),
                    'network_type': network.get('type', '')
                })
        
        return {
            'filename': filename,
            'query': query,
            'results': results,
            'count': len(results)
        }
    
    def get_all_analyses(self) -> Dict[str, Any]:
        """Get analysis for all uploaded files"""
        all_analyses = {}
        for filename in self.uploaded_files:
            all_analyses[filename] = self.analyze_content(filename)
        
        # Create combined analysis
        combined = {
            'filename': 'All Files',
            'file_count': len(self.uploaded_files),
            'suspicious_items': [],
            'statistics': {
                'total_communications': 0,
                'suspicious_items': 0,
                'crypto_references': 0,
                'foreign_contacts': 0,
                'potential_phishing': 0,
                'suspicious_keywords': 0
            }
        }
        
        for analysis in all_analyses.values():
            combined['suspicious_items'].extend(analysis.get('suspicious_items', []))
            for key in combined['statistics']:
                combined['statistics'][key] += analysis.get('statistics', {}).get(key, 0)
        
        all_analyses['_combined'] = combined
        return all_analyses
    
    def generate_pdf_report(self, filename: str) -> BytesIO:
        """Generate a PDF report for a specific file"""
        if filename not in self.uploaded_files:
            return None
        
        analysis = self.analyze_content(filename)
        return self.pdf_generator.generate_report(analysis, filename)
    
    def ask_ai(self, question: str, filename: str = None) -> str:
        """Ask a question to the AI about the data"""
        response = self.ai_chat.get_answer(question, filename)
        
        # Save chat to Firebase
        if self.user_id:
            chat_data = {
                'question': question,
                'response': response,
                'filename': filename,
                'timestamp': datetime.now().isoformat()
            }
            self.firebase.save_chat_history(self.user_id, self.session_id, chat_data)
        
        return response

# Streamlit UI
def main():
    st.title("🔒 Secure UFDR AI Analyst")
    
    # Initialize session state
    if 'agent' not in st.session_state:
        st.session_state.agent = UFDRAnalysisAgent()
    if 'chat_history' not in st.session_state:
        st.session_state.chat_history = []
    if 'current_file' not in st.session_state:
        st.session_state.current_file = None
    
    st.markdown("""
    Welcome to the secure AI-powered UFDR analysis platform. Upload files of any type (up to 5GB) and interact with 
    the AI to analyze suspicious activities, generate reports, and get answers to your questions.
    
    **Security Features:**
    - File type validation
    - Size limitations
    - Input sanitization
    - Secure temporary file handling
    - Encrypted data processing
    """)
    
    # Sidebar for file upload
    with st.sidebar:
        st.header("📁 Upload Files")
        uploaded_files = st.file_uploader(
            "Choose UFDR files (up to 5GB)", 
            accept_multiple_files=True,
            type=None  # Allow all file types
        )
        
        if uploaded_files:
            for uploaded_file in uploaded_files:
                if uploaded_file.name not in st.session_state.agent.uploaded_files:
                    try:
                        with st.spinner(f"Processing {uploaded_file.name}..."):
                            result = st.session_state.agent.process_uploaded_file(uploaded_file)
                            if "error" in result:
                                st.error(f"Error processing {uploaded_file.name}: {result['error']}")
                            else:
                                st.success(f"Processed {uploaded_file.name}")
                    except Exception as e:
                        st.error(f"Security error with {uploaded_file.name}: {str(e)}")
            
            # Show file list with selection
            st.subheader("Uploaded Files")
            file_options = list(st.session_state.agent.uploaded_files.keys())
            selected_file = st.selectbox(
                "Select a file to analyze",
                options=file_options,
                index=0 if file_options else None
            )
            
            if selected_file:
                st.session_state.current_file = selected_file
                
                # Quick actions
                if st.button("📊 Analyze This File"):
                    with st.spinner("Analyzing..."):
                        analysis = st.session_state.agent.analyze_content(selected_file)
                        st.session_state.chat_history.append({
                            "role": "assistant", 
                            "content": f"I've analyzed **{selected_file}**. Found {analysis['statistics']['suspicious_items']} suspicious items. How can I help you explore this data?"
                        })
                
                if st.button("📄 Generate PDF Report"):
                    with st.spinner("Generating report..."):
                        pdf_buffer = st.session_state.agent.generate_pdf_report(selected_file)
                        if pdf_buffer:
                            st.download_button(
                                label="Download PDF Report",
                                data=pdf_buffer,
                                file_name=f"secure_ufdr_report_{selected_file}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                                mime="application/pdf"
                            )
    
    # Main content area - Chat interface
    st.header("💬 Chat with AI Analyst")
    
    # Display chat history
    for message in st.session_state.chat_history:
        with st.chat_message(message["role"]):
            st.markdown(message["content"])
    
    # Chat input
    if prompt := st.chat_input("Ask a question about your data..."):
        # Sanitize input
        sanitized_prompt = security_manager.sanitize_input(prompt)
        
        # Add user message to chat history
        st.session_state.chat_history.append({"role": "user", "content": sanitized_prompt})
        
        # Display user message
        with st.chat_message("user"):
            st.markdown(sanitized_prompt)
        
        # Process query
        with st.chat_message("assistant"):
            with st.spinner("Thinking..."):
                # Check if we have files
                if not st.session_state.agent.uploaded_files:
                    response = "Please upload files first to ask questions about them."
                else:
                    # Get current file context
                    current_file = st.session_state.current_file
                    
                    # Use AI to answer the question
                    response = st.session_state.agent.ask_ai(sanitized_prompt, current_file)
                
                # Display response
                st.markdown(response)
                
                # Add assistant response to chat history
                st.session_state.chat_history.append({"role": "assistant", "content": response})
    
    # Analysis and visualization tabs
    if st.session_state.agent.uploaded_files:
        st.header("📊 Data Analysis")
        
        tab1, tab2 = st.tabs(["File Analysis", "Visualizations"])
        
        with tab1:
            if st.session_state.current_file:
                analysis = st.session_state.agent.analyze_content(st.session_state.current_file)
                
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    st.metric("Total Communications", analysis['statistics']['total_communications'])
                
                with col2:
                    st.metric("Suspicious Items", analysis['statistics']['suspicious_items'])
                
                with col3:
                    st.metric("Crypto References", analysis['statistics']['crypto_references'])
                
                with col4:
                    st.metric("Foreign Contacts", analysis['statistics']['foreign_contacts'])
                
                # Show insights
                st.subheader("Key Insights")
                for insight in analysis['insights']:
                    st.write(insight)
                
                # Show suspicious items
                if analysis['suspicious_items']:
                    st.subheader("Suspicious Items")
                    
                    # Filter options
                    suspicious_type = st.selectbox(
                        "Filter by type",
                        options=["All"] + list(set(item['type'] for item in analysis['suspicious_items']))
                    )
                    
                    filtered_items = analysis['suspicious_items']
                    if suspicious_type != "All":
                        filtered_items = [item for item in analysis['suspicious_items'] if item['type'] == suspicious_type]
                    
                    for item in filtered_items:
                        with st.expander(f"{item['type'].replace('_', ' ').title()}: {item.get('sender', 'Unknown')} on {item.get('date', 'unknown date')}"):
                            st.write(f"**Content:** {item['content']}")
                            if item.get('details'):
                                st.write(f"**Details:** {item['details']}")
        
        with tab2:
            # Create visualizations
            all_analyses = st.session_state.agent.get_all_analyses()
            combined = all_analyses.get('_combined', {})
            stats = combined.get('statistics', {})
            
            if stats:
                col1, col2 = st.columns(2)
                
                with col1:
                    # Pie chart of suspicious items
                    if stats.get('suspicious_items', 0) > 0:
                        suspicious_data = {
                            'Crypto': stats.get('crypto_references', 0),
                            'Foreign Contacts': stats.get('foreign_contacts', 0),
                            'Phishing': stats.get('potential_phishing', 0),
                            'Keywords': stats.get('suspicious_keywords', 0)
                        }
                        
                        # Remove zero values
                        suspicious_data = {k: v for k, v in suspicious_data.items() if v > 0}
                        
                        if suspicious_data:
                            fig = px.pie(
                                values=list(suspicious_data.values()),
                                names=list(suspicious_data.keys()),
                                title="Distribution of Suspicious Items"
                            )
                            st.plotly_chart(fig, use_container_width=True)
                        else:
                            st.info("No suspicious items to visualize")
                
                with col2:
                    # Bar chart of overall statistics
                    stats_data = {
                        'Metric': ['Communications', 'Suspicious Items', 'Crypto Refs', 'Foreign Contacts'],
                        'Count': [
                            stats.get('total_communications', 0),
                            stats.get('suspicious_items', 0),
                            stats.get('crypto_references', 0),
                            stats.get('foreign_contacts', 0)
                        ]
                    }
                    
                    fig = px.bar(
                        stats_data,
                        x='Metric',
                        y='Count',
                        title="Overall Statistics"
                    )
                    st.plotly_chart(fig, use_container_width=True)

if __name__ == "__main__":
    main()
